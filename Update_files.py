import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import os
from win32com import client
from timeit import default_timer as timer

start = timer()
excel = client.DispatchEx("Excel.Application")
excel.Visible = 0

#Main files directory
MainDir = "C:/Users/macie/Pulpit/III_CAT_SPRAWDZENIE_SRUB/20210519_LISTA_WPALEN_python/"
#List name
ListName = "LISTA_WPALEK.xlsx"
#Name of list worksheet
WorkSheetName = "PY"
#Letter of column with names for naming files
Col_lett = "C"
#Range of column with names of files
ColStart = 7
ColEnd = 416
#Template name
TemplateName = "karta_wpalki-08.07.2021.xlsx"
#Name of template worksheet 
TemplateWorkSheetName = "List1"
#Files save / update directory
FilesDir = "C:/Users/macie/Pulpit/III_CAT_SPRAWDZENIE_SRUB/20210519_LISTA_WPALEN_python/WYDRUKI/"
#List of files to create / update
FilesList = []
#Change color marker
MyColor = "E7CE63"
#Map of cells connection between list and template
#TODO #1 Do double translation in excel #2 Map rest of the cells
Map = {
'G1':'AF',
'G34':'AF',
'E2':'AC',  #Installation name
'E35':'AC',  #Installation name
'Q3':'V',  #New pipeline name
'A7':'A',   #Next No acc. to tie-in  point list
'B7':'C',
'C7':'B',
'D7':'F',
'F7':'D',
'I7':'G',
'M7':'J',
'K7':'E',
'N7':'Y',
'O7':'K',
'P7':'AA',
'Q7':'AB',
'R7':'W',
'S7':'X',
'I9':'L',
'H12':'U',
'A15':'L',
'C15':'H',
'D15':'M',
'E15':'O',
'F15':'Q',
'G15':'N',
'H15':'P',
'I15':'E',
'J15':'I',
'L15':'R',
'M15':'S',
'N15':'T',
'D26':'AD',
'E33':'AG',
'K33':'AE'}

#Creating a list of all files name
try:
    source = openpyxl.load_workbook(MainDir + TemplateName)
    wb = openpyxl.load_workbook(MainDir + ListName)
    worksheet = wb[WorkSheetName]
    i = 0
    for z in range(ColStart,ColEnd):
        file = worksheet[f'{Col_lett}{z}'].value
        NewFileName = f'{FilesDir}{file}_between'
        #Check if file already exist
        if os.path.isfile(f'{NewFileName}.xlsx'):
            print(f'{file}_between.xlsx already exist')
            NewFile = openpyxl.load_workbook(f'{NewFileName}.xlsx')
            NewWorksheet = NewFile[TemplateWorkSheetName]
            #Fullfil created files with data from list
            for x,y in Map.items():
                if NewWorksheet[x].value != worksheet[f'{y}{z}'].value:
                    comment = Comment(f'Previous value = {NewWorksheet[x].value}','automatic inspect')
                    NewWorksheet[x].value = worksheet[f'{y}{z}'].value
                    NewWorksheet[x].fill = PatternFill(fgColor=MyColor, fill_type="solid")
                    NewWorksheet[x].comment = comment
            NewFile.save(f'{NewFileName}.xlsx')
            WbPrint = excel.Workbooks.Open(f'{NewFileName}.xlsx')
            WsPrint = WbPrint.Worksheets[TemplateWorkSheetName]
            WbPrint.SaveAs(f"{NewFileName}.pdf",FileFormat=57)
            WbPrint.Close()
            excel.Quit()
            NewFile.close()
        else:
            #Fullfil created files with data from list
            NewWorksheet = source[TemplateWorkSheetName]
            for x,y in Map.items():
                NewWorksheet[x].value = worksheet[f'{y}{z}'].value
            source.save(f'{NewFileName}.xlsx')
            WbPrint = excel.Workbooks.Open(f'{NewFileName}.xlsx')
            WsPrint = WbPrint.Worksheets[TemplateWorkSheetName]
            WbPrint.SaveAs(f"{NewFileName}.pdf",FileFormat=57)
            WbPrint.Close()
            excel.Quit()
        i += 1
        #Close list
    source.close()
    wb.close()
    end = timer()
    print(end-start)
except:
    source.close()
    wb.close()
    excel.Quit()